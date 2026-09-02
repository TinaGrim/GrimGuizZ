from datetime import datetime, timezone
import re
from urllib.parse import urlsplit

from bson import ObjectId
from fastapi import APIRouter, Depends, HTTPException, Query

from ..auth import (
    create_access_token,
    get_current_teacher,
    hash_password,
    verify_password,
)
from ..config import media_url, settings
from ..db import get_db
from ..ratelimit import limit
from ..schemas import (
    AssignRequest,
    ChapterCreate,
    ChapterUpdate,
    PasswordUpdate,
    QuestionCreate,
    QuestionUpdate,
    QuizCreate,
    QuizUpdate,
    SendMessage,
    StudentCreate,
    UsernameUpdate,
    now_iso,
    obj_id,
)

router = APIRouter(prefix="/api/teacher", tags=["teacher"], dependencies=[Depends(get_current_teacher)])
auth_router = APIRouter(prefix="/api/teacher", tags=["teacher-auth"])


# ─── Auth (login excluded from auth dependency) ───────────────────────────────

@auth_router.post(
    "/login",
    dependencies=[Depends(limit("teacher-login", max_calls=10, window_seconds=60))],
)
async def teacher_login(payload: dict):
    db = get_db()
    raw_username = (payload.get("username", "") or "").strip()
    if not raw_username:
        raise HTTPException(status_code=401, detail="Invalid username or password")
    # Case-insensitive lookup so "teacher", "TEACHER", and "Teacher" all
    # resolve to the same account. The exact stored value is used for the
    # password hash comparison (still byte-exact), and the canonical
    # stored username is what we hand back to the client.
    teacher = await db.teachers.find_one(
        {"username": {"$regex": f"^{re.escape(raw_username)}$", "$options": "i"}}
    )
    if teacher is None or not verify_password(payload.get("password", ""), teacher["passwordHash"]):
        raise HTTPException(status_code=401, detail="Invalid username or password")
    token = create_access_token(str(teacher["_id"]))
    return {
        "token": token,
        "teacher": {
            "id": str(teacher["_id"]),
            "username": teacher["username"],
        },
    }


@router.get("/me")
async def teacher_me(teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    t = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    return {"id": str(t["_id"]), "username": t["username"], "displayName": t.get("displayName", "")}


@router.post(
    "/me/username",
    dependencies=[Depends(limit("teacher-username", max_calls=5, window_seconds=60))],
)
async def update_username(
    payload: UsernameUpdate,
    teacher_id: str = Depends(get_current_teacher),
):
    """Change the teacher's own username. Requires the bearer token (no
    extra password check) — the Security panel calls this from a session
    the teacher is already signed in to, and re-prompting for the
    password would be a friction tax without security benefit (the JWT
    itself is the proof of identity)."""
    db = get_db()
    new_username = payload.username.strip()
    if not new_username:
        raise HTTPException(status_code=400, detail="Username cannot be empty")
    # Reject if another teacher already has this exact username (case-
    # insensitive) so the login lookup stays unambiguous.
    existing = await db.teachers.find_one(
        {"username": {"$regex": f"^{re.escape(new_username)}$", "$options": "i"}}
    )
    if existing and str(existing["_id"]) != teacher_id:
        raise HTTPException(status_code=409, detail="That username is already taken")
    await db.teachers.update_one(
        {"_id": ObjectId(teacher_id)},
        {"$set": {"username": new_username}},
    )
    # Mint a fresh token so any downstream code that decodes the username
    # from the JWT (none today, but cheap) stays consistent. The existing
    # token is still valid until it expires.
    return {
        "id": teacher_id,
        "username": new_username,
        "displayName": "",
    }


@router.post(
    "/me/password",
    dependencies=[Depends(limit("teacher-password", max_calls=5, window_seconds=60))],
)
async def update_password(
    payload: PasswordUpdate,
    teacher_id: str = Depends(get_current_teacher),
):
    """Change the teacher's own password. Re-verifies the current password
    before accepting the new one so a stolen session token can't be used
    to lock the account down. Returns a fresh JWT so other tabs pick up
    the new credentials without re-prompting for login."""
    db = get_db()
    teacher = await db.teachers.find_one({"_id": ObjectId(teacher_id)})
    if teacher is None:
        raise HTTPException(status_code=404, detail="Teacher not found")
    if not verify_password(payload.currentPassword, teacher["passwordHash"]):
        raise HTTPException(status_code=401, detail="Current password is incorrect")
    if payload.currentPassword == payload.newPassword:
        raise HTTPException(
            status_code=400,
            detail="New password must be different from the current one",
        )
    await db.teachers.update_one(
        {"_id": ObjectId(teacher_id)},
        {"$set": {"passwordHash": hash_password(payload.newPassword)}},
    )
    return {
        "ok": True,
        "token": create_access_token(teacher_id),
    }


# ─── Chapters ─────────────────────────────────────────────────────────────────

@router.get("/chapters")
async def list_chapters(teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    out = []
    async for c in db.chapters.find().sort("name", 1):
        c["id"] = str(c.pop("_id"))
        out.append(c)
    return out


@router.post("/chapters")
async def create_chapter(payload: ChapterCreate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    doc = {
        "name": payload.name.strip(),
        "description": payload.description.strip(),
        "subject": _infer_subject(payload.name),
        "_id": ObjectId(),
    }
    await db.chapters.insert_one(doc)
    doc["id"] = str(doc.pop("_id"))
    return doc


@router.patch("/chapters/{chapter_id}")
async def update_chapter(chapter_id: str, payload: ChapterUpdate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.chapters.find_one({"_id": ObjectId(chapter_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Chapter not found")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if not updates:
        return {**existing, "id": chapter_id}
    if "name" in updates:
        updates["subject"] = _infer_subject(updates["name"])
    await db.chapters.update_one({"_id": existing["_id"]}, {"$set": updates})
    updated = await db.chapters.find_one({"_id": existing["_id"]})
    updated["id"] = str(updated.pop("_id"))
    return updated


@router.delete("/chapters/{chapter_id}")
async def delete_chapter(chapter_id: str, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.chapters.find_one({"_id": ObjectId(chapter_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Chapter not found")

    # Dependency check — never cascade-delete lessons/quizzes.
    lesson_count = await db.lessons.count_documents({"chapterId": chapter_id})
    if lesson_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete this chapter — {lesson_count} lesson(s) still reference it. Move or remove them first.",
        )
    await db.chapters.delete_one({"_id": existing["_id"]})
    return {"ok": True}


def _infer_subject(name: str) -> str:
    n = (name or "").lower()
    math_words = ["algebra", "geometry", "trigonometry", "statistics", "probability", "functions"]
    physics_words = ["motion", "force", "energy", "work", "electricity", "ohm", "waves", "sound", "heat", "temperature"]
    if any(w in n for w in math_words):
        return "math"
    if any(w in n for w in physics_words):
        return "physics"
    return "other"


# ─── Lessons (lightweight support for chapter/quiz flow) ──────────────────────

@router.get("/lessons")
async def list_lessons(teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    out = []
    async for l in db.lessons.find():
        l["id"] = str(l.pop("_id"))
        out.append(l)
    return out


@router.post("/lessons")
async def create_lesson(payload: dict, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    title = payload.get("title", "").strip()
    chapter_id = payload.get("chapterId", "")
    if not title or not chapter_id:
        raise HTTPException(status_code=400, detail="Lesson title and chapter are required")
    try:
        chapter_oid = ObjectId(chapter_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid chapter id")
    doc = {"title": title, "chapterId": chapter_id, "quizIds": [], "_id": ObjectId()}
    await db.lessons.insert_one(doc)
    await db.chapters.update_one({"_id": chapter_oid}, {"$push": {"lessonIds": str(doc["_id"])}})
    doc["id"] = str(doc.pop("_id"))
    return doc


@router.patch("/lessons/{lesson_id}")
async def update_lesson(lesson_id: str, payload: dict, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    try:
        oid = ObjectId(lesson_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid lesson id")
    existing = await db.lessons.find_one({"_id": oid})
    if existing is None:
        raise HTTPException(status_code=404, detail="Lesson not found")

    updates = {}
    if "title" in payload:
        new_title = (payload.get("title") or "").strip()
        if not new_title:
            raise HTTPException(status_code=400, detail="Lesson title cannot be empty")
        updates["title"] = new_title

    if "chapterId" in payload:
        new_chapter = payload.get("chapterId") or ""
        try:
            new_oid = ObjectId(new_chapter)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid chapter id")
        new_doc = await db.chapters.find_one({"_id": new_oid})
        if new_doc is None:
            raise HTTPException(status_code=404, detail="Target chapter not found")
        old_chapter = existing.get("chapterId")
        if old_chapter and old_chapter != new_chapter:
            try:
                await db.chapters.update_one(
                    {"_id": ObjectId(old_chapter)}, {"$pull": {"lessonIds": lesson_id}}
                )
            except Exception:
                pass
            await db.chapters.update_one(
                {"_id": new_oid}, {"$push": {"lessonIds": lesson_id}}
            )
        updates["chapterId"] = new_chapter

    if updates:
        await db.lessons.update_one({"_id": oid}, {"$set": updates})

    updated = await db.lessons.find_one({"_id": oid})
    updated["id"] = str(updated.pop("_id"))
    return updated


@router.delete("/lessons/{lesson_id}")
async def delete_lesson(lesson_id: str, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    try:
        oid = ObjectId(lesson_id)
    except Exception:
        raise HTTPException(status_code=400, detail="Invalid lesson id")
    existing = await db.lessons.find_one({"_id": oid})
    if existing is None:
        raise HTTPException(status_code=404, detail="Lesson not found")

    # Dependency check — never cascade-delete quizzes.
    quiz_count = existing.get("quizIds") and len(existing["quizIds"]) or 0
    if quiz_count > 0:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete this lesson — {quiz_count} quiz(quizzes) still reference it. Archive the quizzes first.",
        )
    chapter_id = existing.get("chapterId")
    if chapter_id:
        try:
            await db.chapters.update_one(
                {"_id": ObjectId(chapter_id)}, {"$pull": {"lessonIds": lesson_id}}
            )
        except Exception:
            pass
    await db.lessons.delete_one({"_id": oid})
    return {"ok": True}


# ─── Quizzes ─────────────────────────────────────────────────────────────────

@router.get("/quizzes")
async def list_quizzes(teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    out = []
    async for q in db.quizzes.find({"status": {"$ne": "archived"}}):
        if q.get("trollVideoId"):
            q["trollVideoId"] = media_url(q["trollVideoId"])
        q["id"] = str(q.pop("_id"))
        out.append(q)
    return out


@router.post("/quizzes")
async def create_quiz(payload: QuizCreate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    troll_url = await _resolve_asset_reference(payload.trollVideoId, kind="trollVideoId")
    doc = payload.dict()
    doc["_id"] = ObjectId()
    doc["questionPoolIds"] = payload.questionPoolIds or []
    doc["trollVideoId"] = troll_url
    await db.quizzes.insert_one(doc)
    # attach to lesson
    lesson = await db.lessons.find_one({"_id": ObjectId(payload.lessonId)})
    if lesson:
        await db.lessons.update_one({"_id": lesson["_id"]}, {"$push": {"quizIds": str(doc["_id"])}})
    doc["id"] = str(doc.pop("_id"))
    if doc.get("trollVideoId"):
        doc["trollVideoId"] = media_url(doc["trollVideoId"])
    return doc


@router.patch("/quizzes/{quiz_id}")
async def update_quiz(quiz_id: str, payload: QuizUpdate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.quizzes.find_one({"_id": ObjectId(quiz_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Quiz not found")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if "trollVideoId" in updates:
        updates["trollVideoId"] = await _resolve_asset_reference(
            updates["trollVideoId"], kind="trollVideoId"
        )
    if updates:
        await db.quizzes.update_one({"_id": existing["_id"]}, {"$set": updates})
    updated = await db.quizzes.find_one({"_id": existing["_id"]})
    updated["id"] = str(updated.pop("_id"))
    if updated.get("trollVideoId"):
        updated["trollVideoId"] = media_url(updated["trollVideoId"])
    return updated


@router.delete("/quizzes/{quiz_id}")
async def delete_quiz(quiz_id: str, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.quizzes.find_one({"_id": ObjectId(quiz_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Quiz not found")
    # Soft-delete/archive to preserve Attempt history.
    await db.quizzes.update_one({"_id": existing["_id"]}, {"$set": {"status": "archived"}})
    # Remove from lesson quizIds
    await db.lessons.update_many({"quizIds": quiz_id}, {"$pull": {"quizIds": quiz_id}})
    return {"ok": True, "archived": True}


# ─── Questions ────────────────────────────────────────────────────────────────

@router.get("/questions")
async def list_questions(teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    out = []
    async for q in db.questions.find().sort("order", 1):
        if q.get("imageUrl"):
            q["imageUrl"] = media_url(q["imageUrl"])
        if q.get("trollVideoId"):
            q["trollVideoId"] = media_url(q["trollVideoId"])
        q["id"] = str(q.pop("_id"))
        out.append(q)
    return out


async def _resolve_asset_reference(value: str | None, *, kind: str) -> str | None:
    """Normalise an image or troll-video reference to a `/uploads/...` URL.

    Accepts an ObjectId pointing to a record in the `assets` collection, a
    `/uploads/...` URL, or — in prod — an absolute URL that resolves back
    to this backend's configured media origin (`R2_PUBLIC_BASE_URL`, else
    `PUBLIC_BASE_URL`; the asset picker returns absolute URLs when a base
    is configured so the Vercel frontend can load them). Absolute URLs
    are stripped back to the relative `/uploads/...` path for storage so
    file-system ops (delete) keep working. Anything else is rejected so a
    teacher can't embed a third-party tracking pixel, a `javascript:` URL,
    or a path-traversal string into a question. `None` (cleared) and empty
    strings pass through.
    """
    if not value:
        return None
    if value.startswith("/uploads/"):
        return value
    if value.startswith(("http://", "https://")):
        bases = {b for b in (settings.r2_public_base, settings.public_base_url) if b}
        try:
            parts = urlsplit(value)
            if parts.path.startswith("/uploads/") and any(
                parts.netloc == urlsplit(base).netloc for base in bases
            ):
                return parts.path
        except Exception:
            pass
        raise HTTPException(
            status_code=400,
            detail=(
                f"{kind} must be an existing asset id or a "
                f"{'/uploads/' if not bases else (next(iter(bases)) + '/uploads/')} URL"
            ),
        )
    try:
        asset = await get_db().assets.find_one({"_id": ObjectId(value)})
    except Exception:
        asset = None
    if asset is None:
        raise HTTPException(
            status_code=400,
            detail=(
                f"{kind} must be an existing asset id or a /uploads/ URL"
            ),
        )
    return asset.get("url")


@router.post("/questions")
async def create_question(payload: QuestionCreate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    required = payload.options if isinstance(payload.options, list) else []
    if len(required) != 5:
        raise HTTPException(status_code=400, detail="Exactly 5 options are required")
    if any(not o for o in required):
        raise HTTPException(status_code=400, detail="All 5 options must be filled in")
    image_url = await _resolve_asset_reference(payload.imageUrl, kind="imageUrl")
    troll_url = await _resolve_asset_reference(payload.trollVideoId, kind="trollVideoId")
    doc = payload.dict()
    doc["_id"] = ObjectId()
    doc["imageUrl"] = image_url
    doc["trollVideoId"] = troll_url
    await db.questions.insert_one(doc)
    # attach to quiz pool
    await db.quizzes.update_one({"_id": ObjectId(payload.quizId)}, {"$push": {"questionPoolIds": str(doc["_id"])}})
    doc["id"] = str(doc.pop("_id"))
    if doc.get("imageUrl"):
        doc["imageUrl"] = media_url(doc["imageUrl"])
    if doc.get("trollVideoId"):
        doc["trollVideoId"] = media_url(doc["trollVideoId"])
    return doc


@router.patch("/questions/{question_id}")
async def update_question(question_id: str, payload: QuestionUpdate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.questions.find_one({"_id": ObjectId(question_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Question not found")
    updates = {k: v for k, v in payload.dict(exclude_unset=True).items() if v is not None}
    if "options" in updates:
        if len(updates["options"]) != 5 or any(not o for o in updates["options"]):
            raise HTTPException(status_code=400, detail="Exactly 5 non-empty options are required")
    if "imageUrl" in updates:
        updates["imageUrl"] = await _resolve_asset_reference(
            updates["imageUrl"], kind="imageUrl"
        )
    if "trollVideoId" in updates:
        updates["trollVideoId"] = await _resolve_asset_reference(
            updates["trollVideoId"], kind="trollVideoId"
        )
    if updates:
        await db.questions.update_one({"_id": existing["_id"]}, {"$set": updates})
    updated = await db.questions.find_one({"_id": existing["_id"]})
    updated["id"] = str(updated.pop("_id"))
    if updated.get("imageUrl"):
        updated["imageUrl"] = media_url(updated["imageUrl"])
    if updated.get("trollVideoId"):
        updated["trollVideoId"] = media_url(updated["trollVideoId"])
    return updated


@router.delete("/questions/{question_id}")
async def delete_question(question_id: str, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.questions.find_one({"_id": ObjectId(question_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Question not found")
    # Dependency check: is it attached to any non-archived quiz?
    qz = await db.quizzes.find_one({"questionPoolIds": question_id, "status": {"$ne": "archived"}})
    if qz:
        raise HTTPException(
            status_code=409,
            detail=f"Cannot delete — this question is still in the quiz \"{qz.get('title')}\". Remove it from the quiz first.",
        )
    await db.questions.delete_one({"_id": existing["_id"]})
    return {"ok": True}


# ─── Students CRUD ────────────────────────────────────────────────────────────

@router.get("/students")
async def list_students(teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    out = []
    async for s in db.students.find().sort("name", 1):
        s["id"] = str(s.pop("_id"))
        out.append(s)
    return out


@router.post("/students")
async def create_student(payload: StudentCreate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    # Escape the name so a teacher can't create a duplicate-check regex that
    # matches arbitrary students; also avoids Mongo treating backslashes oddly.
    import re
    name_re = re.escape(payload.name.strip())
    existing = await db.students.find_one({"name": {"$regex": f"^{name_re}$", "$options": "i"}})
    if existing:
        raise HTTPException(status_code=409, detail="A student with that name already exists")
    doc = {"name": payload.name.strip(), "createdAt": now_iso(), "assignedQuizIds": [], "_id": ObjectId()}
    await db.students.insert_one(doc)
    doc["id"] = str(doc.pop("_id"))
    return doc


@router.patch("/students/{student_id}")
async def update_student(student_id: str, payload: StudentCreate, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.students.find_one({"_id": ObjectId(student_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Student not found")
    updates = {"name": payload.name.strip()}
    await db.students.update_one({"_id": existing["_id"]}, {"$set": updates})
    updated = await db.students.find_one({"_id": existing["_id"]})
    updated["id"] = str(updated.pop("_id"))
    return updated


@router.delete("/students/{student_id}")
async def delete_student(student_id: str, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    existing = await db.students.find_one({"_id": ObjectId(student_id)})
    if existing is None:
        raise HTTPException(status_code=404, detail="Student not found")
    await db.students.delete_one({"_id": existing["_id"]})
    await db.messages.delete_many({"studentId": student_id})
    return {"ok": True}


@router.post("/assign")
async def assign_quizzes(payload: AssignRequest, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    for sid in payload.studentIds:
        try:
            student = await db.students.find_one({"_id": ObjectId(sid)})
        except Exception:
            continue
        if student is None:
            continue
        assigned = student.get("assignedQuizIds", [])
        if payload.quizId not in assigned:
            await db.students.update_one({"_id": student["_id"]}, {"$push": {"assignedQuizIds": payload.quizId}})
    return {"ok": True}


@router.post("/students/{student_id}/assign")
async def assign_single(student_id: str, payload: dict, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    quiz_id = payload.get("quizId")
    if not quiz_id:
        raise HTTPException(status_code=400, detail="quizId is required")
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    assigned = student.get("assignedQuizIds", [])
    if quiz_id not in assigned:
        await db.students.update_one({"_id": student["_id"]}, {"$push": {"assignedQuizIds": quiz_id}})
    return {"ok": True}


@router.post("/students/{student_id}/unassign")
async def unassign_one(student_id: str, payload: dict, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    quiz_id = payload.get("quizId")
    if not quiz_id:
        raise HTTPException(status_code=400, detail="quizId is required")
    try:
        student = await db.students.find_one({"_id": ObjectId(student_id)})
    except Exception:
        raise HTTPException(status_code=404, detail="Student not found")
    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    await db.students.update_one(
        {"_id": student["_id"]},
        {"$pull": {"assignedQuizIds": quiz_id}},
    )
    return {"ok": True}


# ─── Messages ─────────────────────────────────────────────────────────────────

@router.post("/messages")
async def send_message(payload: SendMessage, teacher_id: str = Depends(get_current_teacher)):
    db = get_db()
    student = await db.students.find_one({"_id": ObjectId(payload.studentId)})
    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")
    doc = {
        "_id": ObjectId(),
        "studentId": payload.studentId,
        "teacherId": teacher_id,
        "text": payload.text.strip(),
        "createdAt": now_iso(),
        "readAt": None,
    }
    await db.messages.insert_one(doc)
    doc["id"] = str(doc.pop("_id"))
    return doc


# ─── Reports (addendum) ─────────────────────────────────────────────────────────

@router.get("/reports/class")
async def class_report(
    range: str = "month",
    teacher_id: str = Depends(get_current_teacher),
):
    """Class-wide aggregate report (addendum §3.2)."""
    from ..reports import build_class_report
    db = get_db()
    return await build_class_report(db, range)


@router.get("/reports/class/export.csv")
async def class_report_csv(
    range: str = "month",
    teacher_id: str = Depends(get_current_teacher),
):
    """Student report — clean CSV with no internal IDs.

    Sections (separated by blank rows):
      1) Header           — class identity + generation timestamp
      2) Students         — one row per student with overall metrics
      3) Student quizzes  — one row per completed attempt (primary block)
      4) Answer trace     — one row per individual question response with
                            options split into 5 columns A–E.
    """
    import csv
    import io
    from datetime import datetime, timezone
    import re

    from ..reports import build_class_report
    db = get_db()

    class_report = await build_class_report(db, range)
    now_iso = datetime.now(timezone.utc).isoformat()

    buf = io.StringIO()
    w = csv.writer(buf)

    def letter(idx: int) -> str:
        return "ABCDE"[idx] if 0 <= idx < 5 else "?"

    # ---- Caches ----
    student_cache: dict[str, dict] = {}

    async def get_student(sid: str) -> dict | None:
        if not sid:
            return None
        if sid in student_cache:
            return student_cache.get(sid)
        try:
            s = await db.students.find_one({"_id": ObjectId(sid)})
        except Exception:
            s = None
        if s:
            s["id"] = str(s["_id"])
        student_cache[sid] = s
        return s

    quiz_cache: dict[str, dict] = {}

    async def get_quiz(qid: str) -> dict | None:
        if not qid:
            return None
        if qid in quiz_cache:
            return quiz_cache.get(qid)
        try:
            q = await db.quizzes.find_one({"_id": ObjectId(qid)})
        except Exception:
            q = None
        if q:
            q["id"] = str(q["_id"])
        quiz_cache[qid] = q
        return q

    question_cache: dict[str, dict] = {}

    async def get_question(qid: str) -> dict | None:
        if not qid:
            return None
        if qid in question_cache:
            return question_cache.get(qid)
        try:
            q = await db.questions.find_one({"_id": ObjectId(qid)})
        except Exception:
            q = None
        if q:
            q["id"] = str(q["_id"])
        question_cache[qid] = q
        return q

    # ---- 1) Header ----
    w.writerow(["QuizZ Student Report"])
    w.writerow(["generated_at", now_iso])
    w.writerow(["range", range])
    w.writerow(["total_students", class_report.get("totalStudents", 0)])
    w.writerow(["active_students", class_report.get("activeStudents", 0)])
    w.writerow([])

    # ---- 2) Students ----
    w.writerow([
        "student_name", "status", "attempts",
        "first_try_correct", "first_try_total", "first_try_rate_percent",
        "overall_percent", "trend", "last_active_at",
    ])
    for s in class_report.get("students", []):
        ft_count = s.get("firstTryCorrectCount", 0) or 0
        ft_total = s.get("firstTryQuestions", 0) or 0
        ft_rate = round((ft_count / ft_total) * 100, 1) if ft_total else 0
        w.writerow([
            s.get("name", ""),
            s.get("status", ""),
            s.get("attemptCount", 0),
            ft_count,
            ft_total,
            ft_rate,
            s.get("overallPercent", 0),
            s.get("trend", ""),
            s.get("lastActiveAt", ""),
        ])
    w.writerow([])

    # ---- Load attempts ----
    since_iso = _since_for_range(range)
    attempts: list[dict] = [a async for a in db.attempts.find({"status": "completed"})]
    if since_iso:
        attempts = [a for a in attempts if (a.get("completedAt") or "") >= since_iso]

    # Stable order: by student name, then completedAt.
    name_for_attempt: dict[str, str] = {}
    for a in attempts:
        s = await get_student(a.get("userId", ""))
        name_for_attempt[str(a["_id"])] = (s or {}).get("name", "")
    attempts.sort(key=lambda a: (
        name_for_attempt.get(str(a["_id"]), "").lower(),
        a.get("completedAt", ""),
    ))

    # ---- 3) Student Quizzes ----
    w.writerow([
        "student_name", "quiz", "attempt_number", "completed_at",
        "score", "total", "score_percent", "total_time_seconds", "wheel_result",
    ])
    attempt_counter: dict[tuple[str, str], int] = {}
    for a in attempts:
        student = await get_student(a.get("userId", ""))
        student_name = (student or {}).get("name", "")
        quiz = await get_quiz(a.get("quizId", "") or "")
        quiz_title = (quiz or {}).get("title", "")

        key = (a.get("userId", ""), a.get("quizId", ""))
        attempt_counter[key] = attempt_counter.get(key, 0) + 1

        score = a.get("score", 0) or 0
        total = a.get("total", 0) or 0
        score_percent = round((score / total) * 100, 1) if total else 0
        total_time = a.get("totalTimeSpentSeconds", 0.0) or 0.0

        w.writerow([
            student_name,
            quiz_title,
            attempt_counter[key],
            a.get("completedAt", ""),
            score,
            total,
            score_percent,
            round(total_time, 2),
            a.get("wheelResult", ""),
        ])
    w.writerow([])

    # ---- 4) Answer Trace ----
    w.writerow([
        "student_name", "quiz", "attempt_number", "completed_at",
        "question_prompt",
        "option_A", "option_B", "option_C", "option_D", "option_E",
        "correct_option", "chosen_option",
        "tries_to_solve", "first_try_correct", "trolled",
        "time_spent_seconds",
    ])
    attempt_counter.clear()
    for a in attempts:
        student = await get_student(a.get("userId", ""))
        student_name = (student or {}).get("name", "")
        quiz = await get_quiz(a.get("quizId", "") or "")
        quiz_title = (quiz or {}).get("title", "")
        completed_at = a.get("completedAt", "")

        key = (a.get("userId", ""), a.get("quizId", ""))
        attempt_counter[key] = attempt_counter.get(key, 0) + 1
        attempt_number = attempt_counter[key]

        for ans in a.get("answers", []) or []:
            q = await get_question(ans.get("questionId", "") or "")
            if not q:
                continue
            options = (q.get("options", []) or []) + ["", "", "", "", ""]
            correct_idx = q.get("correctOptionIndex", -1)
            chosen_idx = ans.get("chosenOptionIndex", -1)
            tries = ans.get("tries", 0) or 0
            ftc = bool(ans.get("firstTryCorrect", False))
            trolled = bool(ans.get("trolled", False))
            ans_time = float(ans.get("timeSpentSeconds", 0.0) or 0.0)

            w.writerow([
                student_name,
                quiz_title,
                attempt_number,
                completed_at,
                q.get("prompt", ""),
                options[0],
                options[1],
                options[2],
                options[3],
                options[4],
                letter(correct_idx) if 0 <= correct_idx < 5 else "",
                letter(chosen_idx) if 0 <= chosen_idx < 5 else "",
                tries,
                "yes" if ftc else "no",
                "yes" if trolled else "no",
                round(ans_time, 2),
            ])

    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="text/csv",
        headers={
            "Content-Disposition": f'attachment; filename="quizZ-student-report-{range}.csv"'
        },
    )


@router.get(
    "/reports/class/export.xlsx",
    dependencies=[Depends(limit("export-class", max_calls=20, window_seconds=60))],
)
async def export_class_matrix_xlsx(
    range_filter: str = Query("month", alias="range"),
    teacher_id: str = Depends(get_current_teacher),
):
    """Class-wide gradebook matrix as .xlsx — addendum excel-export-refactor.

    Sheet 1  Class Overview  — every student × every quiz (latest score)
    Sheet 2  Quiz Detail     — every attempt row (history preserved)
    Sheet 3  Student Summary — compact per-student stats
    """
    import io
    from datetime import datetime as _dt, timezone as _tz
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
    from ..reports import (
        _attempt_percent,
        _first_try_correct_count,
        build_class_matrix,
        build_class_report,
    )

    db = get_db()
    matrix_data = await build_class_matrix(db, range_filter)

    students = matrix_data["students"]
    quiz_cols = matrix_data["quizzes"]
    chapters = matrix_data["chapters"]
    mx = matrix_data["matrix"]
    col_avg = matrix_data["columnAvg"]
    overall_avg = matrix_data["overallAvg"]

    # ── Palette ───────────────────────────────────────────────────────────
    INK       = "1C0F00"
    CREAM     = "FAF1E0"
    CREAM_DARK= "E6D8B8"
    AMBER     = "F0A500"
    AMBER_DARK= "C88A00"
    TEAL      = "0D6E6E"
    TEAL_DARK = "0A4F4F"

    BORDER = Border(
        left=Side("thin", color=CREAM_DARK),
        right=Side("thin", color=CREAM_DARK),
        top=Side("thin", color=CREAM_DARK),
        bottom=Side("thin", color=CREAM_DARK),
    )
    HEADER_FILL  = PatternFill("solid", fgColor=INK)
    HEADER_FONT  = Font(name="Outfit", size=10, bold=True, color=CREAM)
    HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    BODY_FONT    = Font(name="Outfit", size=10, color=INK)
    BOLD_FONT    = Font(name="Outfit", size=10, bold=True, color=INK)
    MUTED_FONT   = Font(name="Outfit", size=9, color="7A6043")
    BAND_FILL    = PatternFill("solid", fgColor=CREAM)
    # Calm teal gradient for conditional formatting (low → high saturation).
    SCORE_FILLS = [
        PatternFill("solid", fgColor="E8F4F4"),  # 0-20  very muted
        PatternFill("solid", fgColor="C5E8E8"),  # 20-40
        PatternFill("solid", fgColor="A0DCDC"),  # 40-60
        PatternFill("solid", fgColor="72CBCB"),  # 60-80
        PatternFill("solid", fgColor="0D6E6E"),  # 80-100
    ]
    SCORE_FONTS = [
        Font(name="Outfit", size=10, color="4A6A6A"),
        Font(name="Outfit", size=10, color="3A5A5A"),
        Font(name="Outfit", size=10, color="2A4A4A"),
        Font(name="Outfit", size=10, bold=True, color="0D6E6E"),
        Font(name="Outfit", size=10, bold=True, color="FFFFFF"),
    ]
    DASH_FILL  = PatternFill("solid", fgColor="F5F5F5")
    DASH_FONT  = Font(name="Outfit", size=10, color="AAAAAA")

    def _band_row(ws, row: int, end_col: int) -> None:
        if row % 2 == 0:
            for col in range(1, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = BAND_FILL

    def _style_score_cell(cell, pct: float | None) -> None:
        if pct is None:
            cell.value = "—"
            cell.fill = DASH_FILL
            cell.font = DASH_FONT
        else:
            cell.value = round(pct, 1)
            cell.number_format = "0.0\"%\""
            idx = min(int(pct / 20), 4)
            cell.fill = SCORE_FILLS[idx]
            cell.font = SCORE_FONTS[idx]
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # ── Filename ──────────────────────────────────────────────────────────
    range_label = {"week": "Last_7d", "month": "Last_30d", "year": "Last_365d"}.get(range_filter, "All")
    gen_date = _dt.now(_tz.utc).strftime("%Y-%m-%d")
    filename = f"QuizZ-ClassReport-{range_label}-{gen_date}.xlsx"

    wb = Workbook()

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 1 — Class Overview Matrix
    # ══════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Class Overview"

    # Row 1: merged chapter bands
    # Row 2: quiz names
    # Col A = student name, rest = quiz columns, final col = Overall Avg
    total_quiz_cols = len(quiz_cols)
    last_col = 1 + total_quiz_cols + 1  # name + quizzes + overall avg

    # Header row 1 — chapter bands
    ws1.cell(row=1, column=1, value="")
    ws1.cell(row=1, column=1).fill = HEADER_FILL
    for ch in chapters:
        start_col = 2 + ch["start"]   # offset by name col
        end_col   = 2 + ch["end"]
        if start_col == end_col:
            cell = ws1.cell(row=1, column=start_col, value=ch["name"])
        else:
            ws1.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
            cell = ws1.cell(row=1, column=start_col, value=ch["name"])
        cell.fill = PatternFill("solid", fgColor=TEAL_DARK)
        cell.font = Font(name="Outfit", size=10, bold=True, color=CREAM)
        cell.alignment = HEADER_ALIGN
        for c in range(start_col, end_col + 1):
            ws1.cell(row=1, column=c).border = BORDER
    # Overall Avg header
    avg_col = 2 + total_quiz_cols
    cell = ws1.cell(row=1, column=avg_col, value="Overall Avg")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER
    # Fill blank cells in row 1
    for c in range(1, last_col + 1):
        ws1.cell(row=1, column=c).border = BORDER

    # Header row 2 — quiz names
    ws1.cell(row=2, column=1, value="Student").font = HEADER_FONT
    ws1.cell(row=2, column=1).fill = HEADER_FILL
    ws1.cell(row=2, column=1).alignment = HEADER_ALIGN
    ws1.cell(row=2, column=1).border = BORDER
    for i, qz in enumerate(quiz_cols):
        col_idx = 2 + i
        cell = ws1.cell(row=2, column=col_idx, value=qz["title"])
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    cell = ws1.cell(row=2, column=avg_col, value="Avg %")
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER
    ws1.row_dimensions[1].height = 24
    ws1.row_dimensions[2].height = 22

    # Student rows
    for si, s in enumerate(students):
        row = 3 + si
        ws1.cell(row=row, column=1, value=s["name"]).font = BOLD_FONT
        ws1.cell(row=row, column=1).border = BORDER
        qdata = mx.get(s["id"], {})
        for qi, qz in enumerate(quiz_cols):
            cell = ws1.cell(row=row, column=2 + qi)
            pct = qdata.get(qz["id"], {}).get("percent") if qz["id"] in qdata else None
            _style_score_cell(cell, pct)
        # Overall avg
        oa = overall_avg.get(s["id"])
        cell = ws1.cell(row=row, column=avg_col)
        _style_score_cell(cell, oa)
        _band_row(ws1, row, last_col)

    # Class average row
    avg_row = 3 + len(students)
    ws1.cell(row=avg_row, column=1, value="Class Average").font = BOLD_FONT
    ws1.cell(row=avg_row, column=1).fill = PatternFill("solid", fgColor=CREAM_DARK)
    ws1.cell(row=avg_row, column=1).border = BORDER
    for qi, qz in enumerate(quiz_cols):
        cell = ws1.cell(row=avg_row, column=2 + qi)
        ca = col_avg.get(qz["id"])
        _style_score_cell(cell, ca if ca else None)
        cell.font = BOLD_FONT
    # Overall class average
    all_oa = [v for v in overall_avg.values()]
    class_oa = round(sum(all_oa) / len(all_oa), 1) if all_oa else 0
    cell = ws1.cell(row=avg_row, column=avg_col)
    _style_score_cell(cell, class_oa)
    cell.font = BOLD_FONT
    for c in range(1, last_col + 1):
        ws1.cell(row=avg_row, column=c).fill = PatternFill("solid", fgColor=CREAM_DARK)

    # Freeze panes: freeze row 2 + col A
    ws1.freeze_panes = "B3"
    # Auto-size columns
    ws1.column_dimensions["A"].width = 22
    for i in range(total_quiz_cols):
        letter = get_column_letter(2 + i)
        ws1.column_dimensions[letter].width = 14
    ws1.column_dimensions[get_column_letter(avg_col)].width = 14

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 2 — Quiz Detail (every attempt)
    # ══════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Quiz Detail")
    detail_headers = [
        "Student", "Quiz", "Chapter", "Lesson", "Date",
        "Score", "Total", "%", "Time (s)", "First-Try OK", "Trolled",
    ]
    for ci, h in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws2.row_dimensions[1].height = 24

    # Gather all attempts in range with resolved metadata.
    range_durations = {"week": 7, "month": 30, "year": 365}
    days = range_durations.get(range_filter, 30)
    cutoff = (_dt.now(_tz.utc) - __import__("datetime").timedelta(days=days)).isoformat()
    attempt_rows: list[dict] = []
    # Cache for quiz/lesson/chapter lookups.
    _q_cache: dict[str, dict | None] = {}
    _l_cache: dict[str, dict | None] = {}
    _c_cache: dict[str, dict | None] = {}

    async def _get_q(qid: str) -> dict | None:
        if qid not in _q_cache:
            try:
                _q_cache[qid] = await db.quizzes.find_one({"_id": __import__("bson").ObjectId(qid)})
            except Exception:
                _q_cache[qid] = None
        return _q_cache[qid]

    async def _get_l(lid: str) -> dict | None:
        if lid not in _l_cache:
            try:
                _l_cache[lid] = await db.lessons.find_one({"_id": __import__("bson").ObjectId(lid)})
            except Exception:
                _l_cache[lid] = None
        return _l_cache[lid]

    async def _get_c(cid: str) -> dict | None:
        if cid not in _c_cache:
            try:
                _c_cache[cid] = await db.chapters.find_one({"_id": __import__("bson").ObjectId(cid)})
            except Exception:
                _c_cache[cid] = None
        return _c_cache[cid]

    # Build a sid→name map for the detail sheet.
    sid_name = {s["id"]: s["name"] for s in students}

    async for a in db.attempts.find(
        {"status": "completed", "completedAt": {"$gte": cutoff}}
    ).sort("completedAt", -1):
        uid = a.get("userId", "")
        qid = a.get("quizId", "")
        if not uid or not qid:
            continue
        quiz = await _get_q(qid)
        lesson = await _get_l(str(quiz.get("lessonId", ""))) if quiz else None
        chapter = await _get_c(str(lesson.get("chapterId", ""))) if lesson else None
        ftc_count, ftc_total_q = _first_try_correct_count(a.get("answers", []))
        first_try_ok = ftc_count > 0 and ftc_total_q > 0
        trolled = any(
            ans.get("trolled", False) for ans in a.get("answers", [])
        )
        pct = _attempt_percent(a) * 100
        attempt_rows.append({
            "studentName": sid_name.get(uid, uid),
            "quizTitle": quiz.get("title", "") if quiz else "",
            "chapterName": chapter.get("name", "") if chapter else "",
            "lessonTitle": lesson.get("title", "") if lesson else "",
            "completedAt": a.get("completedAt", ""),
            "score": a.get("score", 0),
            "total": a.get("total", 0),
            "pct": round(pct, 1),
            "time": round(a.get("totalTimeSpentSeconds", 0) or 0, 1),
            "firstTry": "Yes" if first_try_ok else "No",
            "trolled": "Yes" if trolled else "No",
        })

    for ri, ar in enumerate(attempt_rows, start=2):
        row_data = [
            ar["studentName"], ar["quizTitle"], ar["chapterName"],
            ar["lessonTitle"], ar["completedAt"][:10] if ar["completedAt"] else "",
            ar["score"], ar["total"], ar["pct"], ar["time"],
            ar["firstTry"], ar["trolled"],
        ]
        for ci, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=ri, column=ci, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if ci == 8:  # percent column
                _style_score_cell(cell, val)
            elif ci in (6, 7, 9):
                cell.alignment = Alignment(horizontal="center", vertical="center")
        _band_row(ws2, ri, len(detail_headers))

    # Enable AutoFilter on Sheet 2
    if attempt_rows:
        ws2.auto_filter.ref = f"A1:{get_column_letter(len(detail_headers))}{1 + len(attempt_rows)}"
    ws2.freeze_panes = "A2"
    # Auto-size detail columns
    detail_widths = [20, 22, 22, 22, 12, 8, 8, 8, 10, 12, 10]
    for ci, w in enumerate(detail_widths):
        ws2.column_dimensions[get_column_letter(ci + 1)].width = w

    # ══════════════════════════════════════════════════════════════════════
    # SHEET 3 — Student Summary (compact)
    # ══════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Student Summary")
    summary_headers = [
        "Student", "Quizzes Attempted", "Avg Score %", "Best Score %",
        "First-Try Correct %", "Total Attempts",
    ]
    for ci, h in enumerate(summary_headers, start=1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    ws3.row_dimensions[1].height = 24

    # Compute per-student summary stats from the class report data.
    class_report_data = await build_class_report(db, range_filter)
    cr_students = {s["id"]: s for s in class_report_data.get("students", [])}

    for si, s in enumerate(students):
        row = 2 + si
        cr = cr_students.get(s["id"], {})
        q_attempted = len(mx.get(s["id"], {}))
        oa = overall_avg.get(s["id"], 0)
        best = cr.get("bestScore", 0)
        ftc_rate = cr.get("firstTryCorrectRate", 0) * 100
        att_count = cr.get("attemptCount", 0)
        row_vals = [s["name"], q_attempted, oa, best, round(ftc_rate, 1), att_count]
        for ci, val in enumerate(row_vals, start=1):
            cell = ws3.cell(row=row, column=ci, value=val)
            cell.font = BOLD_FONT if ci == 1 else BODY_FONT
            cell.border = BORDER
            if ci in (3, 4, 5):
                cell.number_format = "0.0\"%\""
                cell.alignment = Alignment(horizontal="center")
        _band_row(ws3, row, len(summary_headers))

    ws3.freeze_panes = "A2"
    summary_widths = [22, 18, 14, 14, 18, 16]
    for ci, w in enumerate(summary_widths):
        ws3.column_dimensions[get_column_letter(ci + 1)].width = w

    # ── Write to buffer and return ────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )



def _since_for_range(range: str) -> str | None:
    from datetime import datetime, timezone, timedelta
    now = datetime.now(timezone.utc)
    if range == "week":
        since = now - timedelta(days=7)
    elif range == "month":
        since = now - timedelta(days=30)
    elif range == "year":
        since = now - timedelta(days=365)
    else:
        return None
    return since.isoformat()


@router.get("/reports/{student_id}")
async def student_report(
    student_id: str,
    range: str = "month",
    teacher_id: str = Depends(get_current_teacher),
):
    """Teacher-side per-student expanded report (addendum §3.1)."""
    from ..reports import build_student_report
    db = get_db()
    base = await build_student_report(db, student_id, range)
    if base is None:
        raise HTTPException(status_code=404, detail="Student not found")

    # Teacher-only enrichment: status flag, time-on-task, wrong-answer patterns,
    # message history.
    from datetime import datetime, timezone
    from ..reports import _attempt_percent, status_flag
    range_durations = {"week": 7, "month": 30, "year": 365}
    student = await db.students.find_one({"_id": ObjectId(student_id)})
    days = range_durations.get(range, 30)
    cutoff = (datetime.now(timezone.utc) - __import__("datetime").timedelta(days=days)).isoformat()
    raw_attempts = []
    async for a in db.attempts.find(
        {"userId": student_id, "status": "completed", "completedAt": {"$gte": cutoff}}
    ).sort("completedAt", -1):
        raw_attempts.append(a)

    percents = [_attempt_percent(a) for a in raw_attempts]
    last3 = percents[:3]
    prior3 = percents[3:6]
    from ..reports import status_flag
    status = status_flag(last3, prior3)

    # Time on task: median per-question seconds for this student, vs class median.
    student_times: list[float] = []
    for a in raw_attempts:
        answers = a.get("answers", [])
        for ans in answers:
            t = float(ans.get("timeSpentSeconds", 0.0) or 0.0)
            if t > 0:
                student_times.append(t)
    student_median = 0.0
    if student_times:
        s = sorted(student_times)
        mid = len(s) // 2
        student_median = (s[mid] if len(s) % 2 else (s[mid - 1] + s[mid]) / 2)

    class_times: list[float] = []
    async for a in db.attempts.find({"status": "completed", "completedAt": {"$gte": cutoff}}):
        for ans in a.get("answers", []):
            t = float(ans.get("timeSpentSeconds", 0.0) or 0.0)
            if t > 0:
                class_times.append(t)
    class_median = 0.0
    if class_times:
        s = sorted(class_times)
        mid = len(s) // 2
        class_median = (s[mid] if len(s) % 2 else (s[mid - 1] + s[mid]) / 2)

    flag_fast = student_median > 0 and class_median > 0 and student_median < class_median * 0.6
    flag_slow = student_median > 0 and class_median > 0 and student_median > class_median * 1.6

    # Wrong-answer patterns (rebuild from raw_attempts since student report
    # doesn't carry them; we re-extract per-lesson answers).
    from collections import defaultdict
    answers_by_lesson: dict[str, list[dict]] = defaultdict(list)
    for a in raw_attempts:
        q = await db.quizzes.find_one({"_id": ObjectId(a.get("quizId") or "")})
        if not q:
            continue
        lesson = await db.lessons.find_one({"_id": ObjectId(q.get("lessonId") or "")})
        if not lesson:
            continue
        lesson_id = str(lesson["_id"])
        for ans in a.get("answers", []):
            ans2 = dict(ans)
            try:
                qd = await db.questions.find_one({"_id": ObjectId(ans.get("questionId"))})
                ans2["prompt"] = qd.get("prompt", "") if qd else ""
            except Exception:
                pass
            ans2["chapterName"] = None
            # we'll enrich chapter after building patterns
            answers_by_lesson[lesson_id].append(ans2)

    async def _lesson_lookup(lid: str) -> dict:
        try:
            l = await db.lessons.find_one({"_id": ObjectId(lid)})
        except Exception:
            l = None
        if not l:
            return {}
        out = {"title": l.get("title"), "chapterName": None}
        try:
            c = await db.chapters.find_one({"_id": ObjectId(l.get("chapterId") or "")})
            if c:
                out["chapterName"] = c.get("name")
        except Exception:
            pass
        return out

    from ..reports import _wrong_answer_patterns
    patterns = _wrong_answer_patterns(
        dict(answers_by_lesson),
        lambda lid: {},
    )
    # Enrich each pattern's chapterName
    for p in patterns:
        linfo = await _lesson_lookup(p["lessonId"])
        p["lessonTitle"] = linfo.get("title")
        p["chapterName"] = linfo.get("chapterName")

    # Message history (newest first)
    messages = []
    async for m in db.messages.find({"studentId": student_id}).sort("createdAt", -1):
        messages.append({
            "id": str(m["_id"]),
            "text": m.get("text", ""),
            "createdAt": m.get("createdAt"),
            "readAt": m.get("readAt"),
        })

    # Last active timestamp
    last_active = None
    if raw_attempts:
        last_active = max((a.get("completedAt") for a in raw_attempts if a.get("completedAt")), default=None)

    base["student"] = {
        "id": student_id,
        "name": student.get("name") if student else None,
        "lastActiveAt": last_active,
        "status": status,
    }
    base["timeOnTask"] = {
        "perQuestionMedianSeconds": round(student_median, 2),
        "classMedianSeconds": round(class_median, 2),
        "flagFast": flag_fast,
        "flagSlow": flag_slow,
    }
    base["wrongAnswerPatterns"] = patterns
    base["messageHistory"] = messages
    return base


@router.get(
    "/reports/{student_id}/export.xlsx",
    dependencies=[Depends(limit("export-student", max_calls=20, window_seconds=60))],
)
async def export_student_report_xlsx(
    student_id: str,
    range_filter: str = Query("month", alias="range"),
    teacher_id: str = Depends(get_current_teacher),
):
    """Per-student report as a multi-sheet .xlsx workbook.

    Sheets:
      1) Summary         — at-a-glance figures for the teacher / parent
      2) Subject Breakdown — Math vs Physics chapter mastery
      3) Lesson Detail    — Chapter → Lesson drill-down
      4) Attempt History  — every completed attempt in the date range_filter
      5) Score Trend      — score-over-time table + embedded line chart
    """
    import io
    from datetime import datetime as _dt, timezone as _tz
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import LineChart, Reference
    from ..reports import _attempt_percent, build_student_report

    db = get_db()

    try:
        student = await db.students.find_one({"_id": ObjectId(student_id)})
    except Exception:
        student = None
    if student is None:
        raise HTTPException(status_code=404, detail="Student not found")

    report = await build_student_report(db, student_id, range_filter)
    if report is None:
        raise HTTPException(status_code=404, detail="Student not found")

    # Build a fresh attempt log scoped to the date range_filter (for Sheet 4 + 5).
    range_durations = {"week": 7, "month": 30, "year": 365}
    days = range_durations.get(range_filter, 30)
    cutoff = (_dt.now(_tz.utc) - __import__("datetime").timedelta(days=days)).isoformat()
    raw_attempts: list[dict] = []
    async for a in db.attempts.find(
        {"userId": student_id, "status": "completed", "completedAt": {"$gte": cutoff}}
    ).sort("completedAt", 1):
        raw_attempts.append(a)

    # Sanitize filename.
    import re as _re
    safe_name = _re.sub(r"[^A-Za-z0-9_-]+", "_", (student.get("name") or "student")).strip("_") or "student"
    range_label = {"week": "Last_7d", "month": "Last_30d", "year": "Last_365d"}.get(range_filter, "All")
    gen_date = _dt.now(_tz.utc).strftime("%Y-%m-%d")
    filename = f"QuizZ-Report-{safe_name}-{range_label}-{gen_date}.xlsx"

    # ---- Palette (calm, matches the on-screen brand) -----------------------
    INK = "1C0F00"
    CREAM = "FAF1E0"
    CREAM_DARK = "E6D8B8"
    AMBER = "F0A500"
    AMBER_DARK = "C88A00"
    TEAL = "0D6E6E"
    TEAL_DARK = "0A4F4F"
    EMBER = "D94F1E"

    BORDER_THIN = Border(
        left=Side(style="thin", color=CREAM_DARK),
        right=Side(style="thin", color=CREAM_DARK),
        top=Side(style="thin", color=CREAM_DARK),
        bottom=Side(style="thin", color=CREAM_DARK),
    )

    HEADER_FILL = PatternFill("solid", fgColor=INK)
    HEADER_FONT = Font(name="Outfit", size=11, bold=True, color=CREAM)
    HEADER_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)

    BAND_FILL = PatternFill("solid", fgColor=CREAM)
    BODY_FONT = Font(name="Outfit", size=10, color=INK)
    BOLD_BODY = Font(name="Outfit", size=10, bold=True, color=INK)
    DISPLAY_FONT = Font(name="Fraunces", size=22, bold=True, color=INK)
    SUBHEAD_FONT = Font(name="Outfit", size=9, bold=True, color=AMBER_DARK, italic=True)
    LABEL_FONT = Font(name="Outfit", size=10, color=TEAL_DARK)
    VALUE_FONT = Font(name="Outfit", size=11, bold=True, color=INK)
    MUTED_FONT = Font(name="Outfit", size=9, color="7A6043")

    def _autosize(ws, min_widths: list[int] | None = None) -> None:
        for col_cells in ws.columns:
            try:
                col_letter = col_cells[0].column_letter
            except Exception:
                continue
            max_len = 0
            for c in col_cells:
                v = c.value
                if v is None:
                    continue
                txt = str(v)
                # Newlines count as 1 char wide in Excel, but we account for them.
                lines = txt.splitlines()
                longest = max((len(line) for line in lines), default=0)
                if longest > max_len:
                    max_len = longest
            width = max(max_len + 2, 12)
            if min_widths:
                idx = ord(col_letter) - ord("A")
                if 0 <= idx < len(min_widths):
                    width = max(width, min_widths[idx])
            ws.column_dimensions[col_letter].width = width

    def _write_header_row(ws, row: int, headers: list[str], freeze_after: bool = True) -> None:
        for col_idx, label in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col_idx, value=label)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = HEADER_ALIGN
            c.border = BORDER_THIN
        ws.row_dimensions[row].height = 28
        if freeze_after:
            ws.freeze_panes = ws.cell(row=row + 1, column=1)

    def _band(ws, row: int, end_col: int) -> None:
        if row % 2 == 0:
            for col in range(1, end_col + 1):
                cell = ws.cell(row=row, column=col)
                if cell.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = BAND_FILL

    def _apply_borders(ws, top: int, left: int, bottom: int, right: int) -> None:
        for r in range(top, bottom + 1):
            for c in range(left, right + 1):
                ws.cell(row=r, column=c).border = BORDER_THIN

    def _trend_arrow(trend: str) -> str:
        # Avoid emoji arrows in spreadsheets — use plain glyphs.
        if trend == "improving":
            return "UP"
        if trend == "declining":
            return "DOWN"
        return "STEADY"

    def _pct(num: float, denom: float) -> float:
        return (num / denom) if denom else 0.0

    def _mastery_text(mastery: str) -> str:
        return mastery or "—"

    wb = Workbook()

    # =================================================================
    # Sheet 1 — Summary
    # =================================================================
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    ws["B2"] = "Student Report"
    ws["B2"].font = Font(name="Fraunces", size=26, bold=True, color=INK)

    ws["B3"] = (student.get("name") or "—").strip() or "—"
    ws["B3"].font = DISPLAY_FONT

    ws["B4"] = f"Range: {range_label.replace('_', ' ').lower()} · Generated {_dt.now(_tz.utc).strftime('%Y-%m-%d %H:%M UTC')}"
    ws["B4"].font = MUTED_FONT

    # Key figures (compact table)
    start_row = 7
    ws.cell(row=start_row, column=2, value="Metric").font = SUBHEAD_FONT
    ws.cell(row=start_row, column=3, value="Value").font = SUBHEAD_FONT
    ws.cell(row=start_row, column=2).fill = PatternFill("solid", fgColor=CREAM)
    ws.cell(row=start_row, column=3).fill = PatternFill("solid", fgColor=CREAM)
    ws.cell(row=start_row, column=2).border = BORDER_THIN
    ws.cell(row=start_row, column=3).border = BORDER_THIN

    overall_pct = float(report.get("overallPercent", 0) or 0)
    first_try_pct = float(report.get("firstTryCorrectRate", 0) or 0) * 100

    # Use the corrected per-report chapter fields (per addendum §1.5). The
    # backend now computes these with a real two-window comparison for
    # most-improved and a consistent attempts-floor for weakest, and ensures
    # the two never resolve to the same chapter.
    weakest_name = report.get("weakestChapterName")
    most_improved = report.get("mostImprovedChapterName")
    most_improved_delta = report.get("mostImprovedDeltaPercent")

    # Trend row: show the actual numbers being compared (addendum §2.1), not
    # just UP/DOWN/STEADY. We split raw_attempts in chronological order into
    # two halves; falls back to the report's coarse trend if there isn't
    # enough data to be meaningful.
    def _trend_text() -> str:
        if len(raw_attempts) < 2:
            return _trend_arrow(report.get("trend", "steady"))
        ordered = sorted(raw_attempts, key=lambda a: a.get("completedAt", "") or "")
        half = len(ordered) // 2
        earlier = ordered[:half]
        later = ordered[half:]
        if not earlier or not later:
            return _trend_arrow(report.get("trend", "steady"))
        earlier_pct = (sum(_attempt_percent(a) for a in earlier) / len(earlier)) * 100
        later_pct = (sum(_attempt_percent(a) for a in later) / len(later)) * 100
        delta = later_pct - earlier_pct
        if abs(delta) < 1.0:
            return "Steady (no movement above 1pt)"
        if delta > 0:
            return f"Up from {round(earlier_pct, 1)}% to {round(later_pct, 1)}% over the selected range"
        return f"Down from {round(earlier_pct, 1)}% to {round(later_pct, 1)}% over the selected range"

    most_improved_display = most_improved or "—"
    if most_improved and most_improved_delta is not None:
        most_improved_display = f"{most_improved} (+{most_improved_delta} pts)"

    rows = [
        ("Quizzes completed", int(report.get("attemptCount", 0) or 0)),
        ("Overall average score", f"{round(overall_pct, 1)}%"),
        ("First-try correct rate", f"{round(first_try_pct, 1)}%"),
        ("Trend", _trend_text()),
        (
            "Weakest topic",
            weakest_name or "—",
        ),
        (
            "Most improved topic",
            most_improved_display,
        ),
    ]

    r = start_row + 1
    for label, value in rows:
        ws.cell(row=r, column=2, value=label).font = LABEL_FONT
        ws.cell(row=r, column=3, value=value).font = VALUE_FONT
        ws.cell(row=r, column=2).border = BORDER_THIN
        ws.cell(row=r, column=3).border = BORDER_THIN
        _band(ws, r, 3)
        r += 1

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 36
    # Summary's print setup (portrait + fit-to-one-page) is reapplied at the
    # bottom of this function, after the global landscape loop, so it isn't
    # silently overwritten.

    # =================================================================
    # Sheet 2 — Subject Breakdown
    # =================================================================
    ws = wb.create_sheet("Subject Breakdown")
    ws.sheet_view.showGridLines = False

    ws["A1"] = "Subject Breakdown"
    ws["A1"].font = Font(name="Fraunces", size=18, bold=True, color=INK)
    ws["A2"] = "Per-chapter mastery, separated by subject."
    ws["A2"].font = MUTED_FONT

    # Group per-chapter data by subject so Sheet 2 can render Math / Physics
    # side-by-side without re-querying.
    by_subject: dict[str, list] = {"math": [], "physics": [], "other": []}
    for ch in report.get("perChapter", []) or []:
        subj = (ch.get("subject") or "other").lower()
        by_subject.setdefault(subj, []).append(ch)
    # Attach to the worksheet so the helper can read it.
    ws._subject_data = by_subject  # type: ignore[attr-defined]

    _write_subject_table(ws, 4, "Math", "math")
    _write_subject_table(ws, 4, "Physics", "physics", start_col=6)

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 4
    ws.column_dimensions["G"].width = 26
    ws.column_dimensions["H"].width = 14
    ws.column_dimensions["I"].width = 14
    ws.column_dimensions["J"].width = 14

    # =================================================================
    # Sheet 3 — Lesson Detail
    # =================================================================
    ws = wb.create_sheet("Lesson Detail")
    ws.sheet_view.showGridLines = False
    headers = [
        "Chapter",
        "Lesson",
        "Subject",
        "Attempts",
        "Avg Score",
        "First-try Rate",
        "Mastery",
        "Trend",
        "Avg Time / Question (s)",
    ]
    _write_header_row(ws, 1, headers)
    per_lesson = list(report.get("perLesson", []) or [])
    per_lesson.sort(
        key=lambda l: (
            (l.get("chapterName") or "").lower(),
            (l.get("lessonTitle") or "").lower(),
        )
    )
    r = 2
    for l in per_lesson:
        ws.cell(row=r, column=1, value=l.get("chapterName", "—")).font = BODY_FONT
        ws.cell(row=r, column=2, value=l.get("lessonTitle", "—")).font = BODY_FONT
        ws.cell(row=r, column=3, value=(l.get("subject") or "—").title()).font = BODY_FONT
        ws.cell(row=r, column=4, value=int(l.get("attempts", 0) or 0)).font = BODY_FONT
        # Avg score (percent)
        avg_score_pct = (float(l.get("avgScore", 0) or 0)) * 100
        score_cell = ws.cell(row=r, column=5, value=round(avg_score_pct, 1))
        score_cell.font = BODY_FONT
        score_cell.number_format = "0.0\"%\""
        # First-try rate
        ft_rate = float(l.get("firstTryCorrectRate", 0) or 0) * 100
        ft_cell = ws.cell(row=r, column=6, value=round(ft_rate, 1))
        ft_cell.font = BODY_FONT
        ft_cell.number_format = "0.0\"%\""
        # Mastery
        ws.cell(row=r, column=7, value=_mastery_text(l.get("mastery", ""))).font = BOLD_BODY
        # Trend
        ws.cell(row=r, column=8, value=_trend_arrow(l.get("trend", "steady"))).font = BODY_FONT
        # Avg time per question
        avg_t = float(l.get("avgTimePerQuestionSeconds", 0) or 0)
        t_cell = ws.cell(row=r, column=9, value=round(avg_t, 1))
        t_cell.font = BODY_FONT
        t_cell.number_format = "0.0"
        _band(ws, r, len(headers))
        r += 1
    _apply_borders(ws, 1, 1, max(r - 1, 1), len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(r - 1, 1)}"
    _autosize(ws, [22, 32, 12, 12, 14, 18, 18, 12, 22])

    # =================================================================
    # Sheet 4 — Attempt History
    # =================================================================
    ws = wb.create_sheet("Attempt History")
    ws.sheet_view.showGridLines = False
    headers = [
        "Completed",
        "Quiz",
        "Chapter",
        "Lesson",
        "Score",
        "Total",
        "Score %",
        "Time Spent (s)",
        "Wheel Result",
        "First-try Correct",
        "Status",
    ]
    _write_header_row(ws, 1, headers)

    # Resolve chapter/lesson names for each attempt up-front.
    quiz_cache: dict[str, dict] = {}
    lesson_cache: dict[str, dict] = {}
    chapter_cache: dict[str, dict] = {}

    async def _quiz(qid: str) -> dict:
        if qid in quiz_cache:
            return quiz_cache[qid]
        try:
            q = await db.quizzes.find_one({"_id": ObjectId(qid)})
        except Exception:
            q = None
        out = q or {}
        quiz_cache[qid] = out
        return out

    async def _lesson(lid: str) -> dict:
        if lid in lesson_cache:
            return lesson_cache[lid]
        try:
            l = await db.lessons.find_one({"_id": ObjectId(lid)})
        except Exception:
            l = None
        out = l or {}
        lesson_cache[lid] = out
        return out

    async def _chapter(cid: str) -> dict:
        if cid in chapter_cache:
            return chapter_cache[cid]
        try:
            c = await db.chapters.find_one({"_id": ObjectId(cid)})
        except Exception:
            c = None
        out = c or {}
        chapter_cache[cid] = out
        return out

    # Sort newest first.
    raw_attempts.sort(key=lambda a: a.get("completedAt", ""), reverse=True)

    r = 2
    for a in raw_attempts:
        q = await _quiz(a.get("quizId") or "")
        lesson = await _lesson(q.get("lessonId", "") if q else "")
        chapter = await _chapter(lesson.get("chapterId", "") if lesson else "")
        ws.cell(row=r, column=1, value=a.get("completedAt", "")).font = BODY_FONT
        ws.cell(row=r, column=2, value=q.get("title", "—")).font = BODY_FONT
        ws.cell(row=r, column=3, value=chapter.get("name", "—")).font = BODY_FONT
        ws.cell(row=r, column=4, value=lesson.get("title", "—")).font = BODY_FONT
        score = int(a.get("score", 0) or 0)
        total = int(a.get("total", 0) or 0)
        ws.cell(row=r, column=5, value=score).font = BODY_FONT
        ws.cell(row=r, column=6, value=total).font = BODY_FONT
        pct_val = round(_pct(score, total) * 100, 1)
        pct_cell = ws.cell(row=r, column=7, value=pct_val)
        pct_cell.font = BODY_FONT
        pct_cell.number_format = "0.0\"%\""
        t_val = float(a.get("totalTimeSpentSeconds", 0.0) or 0.0)
        t_cell = ws.cell(row=r, column=8, value=round(t_val, 1))
        t_cell.font = BODY_FONT
        t_cell.number_format = "0.0"
        ws.cell(row=r, column=9, value=a.get("wheelResult", "")).font = BODY_FONT
        # First-try correct count for the attempt
        ft = 0
        for ans in a.get("answers", []) or []:
            if ans.get("firstTryCorrect"):
                ft += 1
        ws.cell(row=r, column=10, value=ft).font = BODY_FONT
        ws.cell(row=r, column=11, value=a.get("status", "")).font = BODY_FONT
        _band(ws, r, len(headers))
        r += 1
    _apply_borders(ws, 1, 1, max(r - 1, 1), len(headers))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(r - 1, 1)}"
    _autosize(ws, [22, 28, 22, 26, 8, 8, 12, 18, 14, 16, 12])

    # =================================================================
    # Sheet 5 — Score Trend
    # =================================================================
    ws = wb.create_sheet("Score Trend")
    ws.sheet_view.showGridLines = False
    headers = ["Date", "Score %", "Quiz"]
    _write_header_row(ws, 1, headers)

    trend_rows = sorted(raw_attempts, key=lambda a: a.get("completedAt", ""))
    r = 2
    for a in trend_rows:
        q = await _quiz(a.get("quizId") or "")
        score = int(a.get("score", 0) or 0)
        total = int(a.get("total", 0) or 0)
        pct_val = round(_pct(score, total) * 100, 1)
        date_cell = ws.cell(row=r, column=1, value=a.get("completedAt", ""))
        date_cell.font = BODY_FONT
        date_cell.number_format = "yyyy-mm-dd hh:mm"
        pct_cell = ws.cell(row=r, column=2, value=pct_val)
        pct_cell.font = BODY_FONT
        pct_cell.number_format = "0.0\"%\""
        ws.cell(row=r, column=3, value=q.get("title", "—")).font = BODY_FONT
        _band(ws, r, len(headers))
        r += 1
    last_row = max(r - 1, 1)
    _apply_borders(ws, 1, 1, last_row, len(headers))
    _autosize(ws, [22, 14, 32])

    # Embed a native line chart if we have ≥2 data points.
    if last_row >= 3:
        chart = LineChart()
        chart.title = "Score Trend"
        chart.y_axis.title = "Score %"
        chart.x_axis.title = "Attempt"
        chart.height = 9
        chart.width = 18
        data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=last_row)
        cats = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=last_row)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        # Style the line using the brand teal — avoid red/yellow/green scales.
        try:
            from openpyxl.chart.shapes import GraphicalProperties
            from openpyxl.drawing.line import LineProperties
            from openpyxl.drawing.fill import ColorChoice
            for series in chart.series:
                series.graphicalProperties = GraphicalProperties(
                    solidFill=TEAL,
                )
                series.graphicalProperties.line = LineProperties(
                    solidFill=TEAL,
                    w=22000,
                )
                # Round marker dots
                from openpyxl.chart.marker import Marker
                series.marker = Marker(symbol="circle", size=6)
                series.marker.graphicalProperties = GraphicalProperties(solidFill=TEAL)
        except Exception:
            pass
        ws.add_chart(chart, "E2")

    # ---- Print-friendly defaults for the whole workbook --------------------
    for sheet in wb.worksheets:
        sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
        sheet.page_setup.fitToWidth = 1
        sheet.page_setup.fitToHeight = 0
        sheet.sheet_properties.pageSetUpPr.fitToPage = True
        sheet.page_margins.left = 0.4
        sheet.page_margins.right = 0.4
        sheet.page_margins.top = 0.5
        sheet.page_margins.bottom = 0.5

    # Summary gets portrait + fit-to-one-page so it actually prints to a
    # single sheet of paper (per addendum §3 "print-friendly"). These settings
    # are reapplied here, after the global landscape loop above, so they win.
    if "Summary" in wb.sheetnames:
        summary_ws = wb["Summary"]
        summary_ws.page_setup.orientation = summary_ws.ORIENTATION_PORTRAIT
        summary_ws.page_setup.fitToWidth = 1
        summary_ws.page_setup.fitToHeight = 1
        summary_ws.print_options.horizontalCentered = True
        summary_ws.print_area = "A1:C20"
        summary_ws.page_margins.left = 0.5
        summary_ws.page_margins.right = 0.5
        summary_ws.page_margins.top = 0.5
        summary_ws.page_margins.bottom = 0.5

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    from fastapi.responses import Response
    return Response(
        content=buf.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"'
        },
    )


def _write_subject_table(ws, start_row: int, label: str, subject: str, start_col: int = 2) -> None:
    """Helper used by Sheet 2 — writes a Math/Physics sub-table.

    Defined at module scope (not nested inside the endpoint) so the inner
    closure can find it; openpyxl Worksheet objects need to be passed
    through explicitly.
    """
    # Local import to avoid circular issues at module load time.
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    INK = "1C0F00"
    CREAM = "FAF1E0"
    CREAM_DARK = "E6D8B8"
    AMBER_DARK = "C88A00"
    TEAL_DARK = "0A4F4F"
    HEADER_FILL = PatternFill("solid", fgColor=INK)
    HEADER_FONT = Font(name="Outfit", size=11, bold=True, color=CREAM)
    SUBHEAD_FONT = Font(name="Outfit", size=10, bold=True, color=AMBER_DARK)
    BODY_FONT = Font(name="Outfit", size=10, color=INK)
    BAND_FILL = PatternFill("solid", fgColor=CREAM)
    BORDER_THIN = Border(
        left=Side(style="thin", color=CREAM_DARK),
        right=Side(style="thin", color=CREAM_DARK),
        top=Side(style="thin", color=CREAM_DARK),
        bottom=Side(style="thin", color=CREAM_DARK),
    )

    # Resolve the per-chapter data for this subject from the report via a
    # stashed global. We populate _SUBJECT_DATA on the workbook before calling.
    sub_data = getattr(ws, "_subject_data", {}).get(subject, [])

    title_cell = ws.cell(row=start_row, column=start_col, value=label)
    title_cell.font = SUBHEAD_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    headers = ["Chapter", "Avg Score", "Attempts", "First-try Rate"]
    header_row = start_row + 1
    for i, h in enumerate(headers):
        c = ws.cell(row=header_row, column=start_col + i, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = BORDER_THIN
    ws.row_dimensions[header_row].height = 26

    r = header_row + 1
    if not sub_data:
        ws.cell(row=r, column=start_col, value="No data for this subject.").font = BODY_FONT
        return
    for idx, c in enumerate(sub_data):
        row = r + idx
        ws.cell(row=row, column=start_col, value=c.get("chapterName", "—")).font = BODY_FONT
        score_pct = round(float(c.get("percent", 0) or 0), 1)
        score_cell = ws.cell(row=row, column=start_col + 1, value=score_pct)
        score_cell.font = BODY_FONT
        score_cell.number_format = "0.0\"%\""
        ws.cell(row=row, column=start_col + 2, value=int(c.get("attempts", 0) or 0)).font = BODY_FONT
        ft_pct = round(float(c.get("firstTryCorrectRate", 0) or 0) * 100, 1)
        ft_cell = ws.cell(row=row, column=start_col + 3, value=ft_pct)
        ft_cell.font = BODY_FONT
        ft_cell.number_format = "0.0\"%\""
        # Banding + borders
        if idx % 2 == 1:
            for offset in range(4):
                cell = ws.cell(row=row, column=start_col + offset)
                if cell.fill.fgColor.rgb in (None, "00000000", "FFFFFFFF"):
                    cell.fill = BAND_FILL
        for offset in range(4):
            ws.cell(row=row, column=start_col + offset).border = BORDER_THIN


@router.get("/reports/{student_id}/attempts/{attempt_id}")
async def student_attempt_detail(
    student_id: str,
    attempt_id: str,
    teacher_id: str = Depends(get_current_teacher),
):
    """Full per-question breakdown for one specific attempt (addendum §3.1)."""
    db = get_db()
    try:
        attempt = await db.attempts.find_one({"_id": ObjectId(attempt_id), "userId": student_id})
    except Exception:
        raise HTTPException(status_code=404, detail="Attempt not found")
    if attempt is None:
        raise HTTPException(status_code=404, detail="Attempt not found")

    served = attempt.get("questionsServed", [])
    breakdown = []
    for sq in served:
        q = await db.questions.find_one({"_id": ObjectId(sq["questionId"])})
        ans = next(
            (a for a in attempt.get("answers", []) if a.get("questionId") == sq["questionId"]),
            None,
        )
        breakdown.append({
            "questionId": sq["questionId"],
            "prompt": sq.get("prompt"),
            "imageUrl": media_url(sq.get("imageUrl")),
            "options": sq.get("options", []),
            "correctOptionIndex": q["correctOptionIndex"] if q else -1,
            "chosenOptionIndex": ans.get("chosenOptionIndex") if ans else None,
            "chosenOptionsHistory": ans.get("chosenOptionsHistory", []) if ans else [],
            "tries": ans.get("tries", 0) if ans else 0,
            "firstTryCorrect": bool(ans and ans.get("firstTryCorrect")),
            "correct": bool(ans and ans.get("correct")),
            "trolled": bool(ans and ans.get("trolled")),
            "timeSpentSeconds": float(ans.get("timeSpentSeconds", 0.0) or 0.0) if ans else 0.0,
        })
    return {
        "id": str(attempt["_id"]),
        "studentId": student_id,
        "quizId": attempt.get("quizId"),
        "score": attempt.get("score", 0),
        "total": attempt.get("total", 0),
        "wheelResult": attempt.get("wheelResult"),
        "completedAt": attempt.get("completedAt"),
        "totalTimeSpentSeconds": float(attempt.get("totalTimeSpentSeconds", 0.0) or 0.0),
        "breakdown": breakdown,
    }
